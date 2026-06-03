"""
Gera relatórios de faturamento por laboratório:
  - Arquivo_Diretoria.xlsx (visão completa + ranking)
  - Arquivo_Cliente_Danone.xlsx (dados mascarados para cliente)

Uso:
  python gerar_relatorios_laboratorios.py
  python gerar_relatorios_laboratorios.py "caminho\\base.xlsx"
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from danone.config import BASE_DIR, PASTA_SAIDAS, PLANILHA_FONTE
ARQUIVO_ENTRADA_PADRAO = PLANILHA_FONTE
SAIDA_DIRETORIA = PASTA_SAIDAS / "Arquivo_Diretoria.xlsx"
SAIDA_CLIENTE = PASTA_SAIDAS / "Arquivo_Cliente_Danone.xlsx"

COL_LAB = "Laboratório"
COL_FAT25 = "Faturamento 2025"
COL_FAT26 = "Faturamento 2026"
COL_CRESC = "Crescimento % YoY"
COL_MS = "Market Share %"
COL_RANK = "Ranking"

MARCADORES_PRODUTO = (
    " PO ",
    " G X ",
    " LAT",
    " KIT ",
    " JR ",
    " POLAP",
    " SUSTAIN",
    " TRIDENT ",
    " LACTA ",
    " APTAMIL",
    " APTANUTRI",
    " MILNUTRI",
    " PREGOMIN",
)
FATURAMENTO_MINIMO_LAB = 10_000_000  # R$ 10 mi — ignora linhas de produto residual
NOMES_EXCLUIR = (
    "total geral",
    "total",
    "subtotal",
    "grand total",
    "(blank)",
)

FMT_MOEDA = '"R$" #.##0,00'
FMT_PCT = "0,00%"

HEADER_FILL = PatternFill("solid", fgColor="1A2B4A")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Segoe UI", size=11)
DANONE_FILL = PatternFill("solid", fgColor="E8F5E9")
MASK_FILL = PatternFill("solid", fgColor="F8F9FA")
THIN_BORDER = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)


def _parse_moeda(valor) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto in ("-", "nan", "None"):
        return None
    texto = texto.replace("R$", "").replace("\xa0", " ").strip()
    if not texto or texto == "-":
        return None
    negativo = texto.startswith("(") and texto.endswith(")")
    texto = texto.strip("()").replace(" ", "")
    texto = texto.replace(".", "").replace(",", ".")
    try:
        n = float(texto)
        return -n if negativo else n
    except ValueError:
        return None


def _parse_percentual(valor) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        v = float(valor)
        return v / 100 if abs(v) > 1.5 else v
    texto = str(valor).strip().replace("%", "").replace(" ", "")
    if not texto or texto in ("-", "nan"):
        return None
    texto = texto.replace(",", ".")
    try:
        return float(texto) / 100.0
    except ValueError:
        return None


def _eh_laboratorio(nome: str) -> bool:
    if not nome or not isinstance(nome, str):
        return False
    n = nome.strip()
    if not n or n.upper() in ("LABORATORIO", "LABORATÓRIO", "NAN"):
        return False
    if n.strip().lower() in NOMES_EXCLUIR or n.strip().lower().startswith("total "):
        return False
    upper = n.upper()
    if any(m in upper for m in MARCADORES_PRODUTO):
        return False
    if re.search(r"\d+\.?\d*\s*G\b", upper):
        return False
    if re.search(r"\sX\s*\d", upper):
        return False
    if re.search(r"\d+\.?\d*\s*KG", upper):
        return False
    if len(n.split()) > 6:
        return False
    return True


def _montar_dataframe(labs, fat25, fat26, cresc) -> pd.DataFrame:
    df = pd.DataFrame(
        {
            COL_LAB: labs,
            COL_FAT25: fat25,
            COL_FAT26: fat26,
            COL_CRESC: cresc,
        }
    )
    df = df.dropna(subset=[COL_LAB])
    df = df[df[COL_LAB].astype(str).str.strip() != ""]
    df = df.drop_duplicates(subset=[COL_LAB], keep="first")
    df[COL_FAT25] = pd.to_numeric(df[COL_FAT25], errors="coerce")
    df[COL_FAT26] = pd.to_numeric(df[COL_FAT26], errors="coerce")
    df[COL_CRESC] = pd.to_numeric(df[COL_CRESC], errors="coerce")
    df = df[df[COL_FAT26].notna() & (df[COL_FAT26] >= FATURAMENTO_MINIMO_LAB)]
    df = df.sort_values(COL_FAT26, ascending=False).reset_index(drop=True)
    df[COL_RANK] = range(1, len(df) + 1)
    return df[
        [COL_RANK, COL_LAB, COL_FAT25, COL_FAT26, COL_CRESC]
    ]


def _extrair_de_planilha_resumo(caminho: Path) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(caminho, sheet_name="Resumo Executivo")
    except (ValueError, FileNotFoundError):
        return None
    cols = {c.strip().lower(): c for c in df.columns.astype(str)}
    lab_col = next((cols[k] for k in cols if "laborat" in k), None)
    f25 = next((cols[k] for k in cols if "2025" in k), None)
    f26 = next((cols[k] for k in cols if "2026" in k), None)
    cresc = next((cols[k] for k in cols if "cresc" in k or "%" in k), None)
    if not all([lab_col, f25, f26]):
        return None
    labs, v25, v26, vp = [], [], [], []
    for _, row in df.iterrows():
        lab = row.get(lab_col)
        if not _eh_laboratorio(str(lab) if pd.notna(lab) else ""):
            continue
        labs.append(str(lab).strip())
        v25.append(_parse_moeda(row.get(f25)))
        v26.append(_parse_moeda(row.get(f26)))
        vp.append(_parse_percentual(row.get(cresc)) if cresc else None)
    if not labs:
        return None
    return _montar_dataframe(labs, v25, v26, vp)


def _extrair_de_bloco_bruto(df: pd.DataFrame) -> pd.DataFrame | None:
    """Varre colunas iniciais (exportação de tabela dinâmica / CSV)."""
    if df.empty:
        return None
    col_lab = df.columns[0]
    col_25 = df.columns[1] if len(df.columns) > 1 else None
    col_26 = df.columns[2] if len(df.columns) > 2 else None
    col_c = df.columns[3] if len(df.columns) > 3 else None
    if col_25 is None or col_26 is None:
        return None

    labs, v25, v26, vp = [], [], [], []
    for _, row in df.iterrows():
        lab = row.get(col_lab)
        if not _eh_laboratorio(str(lab) if pd.notna(lab) else ""):
            continue
        f25 = _parse_moeda(row.get(col_25))
        f26 = _parse_moeda(row.get(col_26))
        if f26 is None or f26 <= 0:
            continue
        labs.append(str(lab).strip())
        v25.append(f25)
        v26.append(f26)
        vp.append(_parse_percentual(row.get(col_c)) if col_c else None)

    if not labs:
        return None
    return _montar_dataframe(labs, v25, v26, vp)


def _varrer_planilha_bruta(df: pd.DataFrame) -> pd.DataFrame | None:
    """Percorre toda a planilha (vários blocos de tabela dinâmica)."""
    labs, v25, v26, vp = [], [], [], []
    ncols = min(4, df.shape[1])
    for i in range(len(df)):
        lab = df.iloc[i, 0]
        if not _eh_laboratorio(str(lab) if pd.notna(lab) else ""):
            continue
        f25 = _parse_moeda(df.iloc[i, 1]) if ncols > 1 else None
        f26 = _parse_moeda(df.iloc[i, 2]) if ncols > 2 else None
        if f26 is None or f26 < FATURAMENTO_MINIMO_LAB:
            continue
        pct = _parse_percentual(df.iloc[i, 3]) if ncols > 3 else None
        nome = str(lab).strip()
        if nome in labs:
            continue
        labs.append(nome)
        v25.append(f25)
        v26.append(f26)
        vp.append(pct)
    if not labs:
        return None
    return _montar_dataframe(labs, v25, v26, vp)


def _extrair_de_dados_completos(caminho: Path) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(caminho, sheet_name="Dados Completos", header=None)
    except (ValueError, FileNotFoundError):
        return None
    return _varrer_planilha_bruta(df)


def _extrair_de_csv(caminho: Path) -> pd.DataFrame | None:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            raw = pd.read_csv(caminho, sep=";", header=None, encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return None
    return _extrair_de_dados_completos_from_raw(raw)


def _extrair_de_dados_completos_from_raw(df: pd.DataFrame) -> pd.DataFrame | None:
    return _varrer_planilha_bruta(df)


def ler_base_laboratorios(caminho: Path) -> pd.DataFrame:
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    if caminho.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        if caminho.name.lower().startswith("estudo_danone") or caminho.name.lower().startswith("maio 2026"):
            from danone import carregar_estudo

            dados = carregar_estudo(caminho)
            labs = dados.laboratorios_ranking
            df = pd.DataFrame(
                {
                    COL_LAB: [l.nome for l in labs],
                    COL_FAT25: [l.fat_2025 for l in labs],
                    COL_FAT26: [l.fat_2026 for l in labs],
                    COL_CRESC: [l.crescimento for l in labs],
                    COL_MS: [l.market_share for l in labs],
                }
            )
            df = df.sort_values(COL_FAT26, ascending=False).reset_index(drop=True)
            df[COL_RANK] = range(1, len(df) + 1)
            return df[[COL_RANK, COL_LAB, COL_FAT25, COL_FAT26, COL_CRESC, COL_MS]]

        df = _extrair_de_dados_completos(caminho)
        resumo = _extrair_de_planilha_resumo(caminho)
        if df is not None and resumo is not None:
            combinado = pd.concat([df, resumo], ignore_index=True)
            combinado = combinado.drop_duplicates(subset=[COL_LAB], keep="first")
            combinado = combinado.sort_values(COL_FAT26, ascending=False).reset_index(drop=True)
            combinado[COL_RANK] = range(1, len(combinado) + 1)
            return combinado[[COL_RANK, COL_LAB, COL_FAT25, COL_FAT26, COL_CRESC]]
        if df is not None and len(df) >= 1:
            return df
        if resumo is not None and len(resumo) > 0:
            return resumo

    if caminho.suffix.lower() == ".csv":
        df = _extrair_de_csv(caminho)
        if df is not None:
            return df

    raise ValueError(
        f"Não foi possível extrair laboratórios de {caminho}. "
        "Verifique se há colunas Laboratório, Faturamento 2025/2026."
    )


def _eh_danone(nome: str) -> bool:
    return "danone" in str(nome).lower()


def preparar_visao_danone(df: pd.DataFrame) -> pd.DataFrame:
    """Mascara concorrentes; mantém ranking real e dados Danone."""
    saida = df.copy()
    mascarados = 0
    nomes_mascara: dict[str, str] = {}

    for idx, row in saida.iterrows():
        lab = row[COL_LAB]
        if _eh_danone(lab):
            continue
        mascarados += 1
        letra = chr(ord("A") + mascarados - 1)
        nomes_mascara[lab] = f"Concorrente {letra}"
        saida.at[idx, COL_LAB] = nomes_mascara[lab]
        saida.at[idx, COL_FAT25] = None
        saida.at[idx, COL_FAT26] = None
        saida.at[idx, COL_CRESC] = None
        saida.at[idx, COL_MS] = None

    return saida


def _aplicar_formato_excel(caminho: Path, mascarado: bool = False) -> None:
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Faturamento por Laboratório"

    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    larguras = {1: 10, 2: 32, 3: 22, 4: 22, 5: 18, 6: 16}
    for col_idx, width in larguras.items():
        if col_idx <= max_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for row in range(2, max_row + 1):
        lab_cell = ws.cell(row=row, column=2)
        is_danone = _eh_danone(str(lab_cell.value or ""))
        row_fill = DANONE_FILL if (mascarado and is_danone) else None

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if row_fill and col <= 2:
                cell.fill = row_fill
            if mascarado and not is_danone and col >= 2:
                cell.fill = MASK_FILL

            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:
                cell.alignment = Alignment(horizontal="left")
            elif col in (3, 4):
                if cell.value is not None and cell.value != "":
                    cell.number_format = FMT_MOEDA
                cell.alignment = Alignment(horizontal="right")
            elif col in (5, 6):
                if cell.value is not None and cell.value != "":
                    cell.number_format = FMT_PCT
                cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 22
    wb.save(caminho)


def exportar_excel(df: pd.DataFrame, caminho: Path, mascarado: bool = False) -> None:
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(caminho, index=False, sheet_name="Faturamento por Laboratório")
    _aplicar_formato_excel(caminho, mascarado=mascarado)


def gerar_relatorios(
    entrada: Path | None = None,
    saida_diretoria: Path = SAIDA_DIRETORIA,
    saida_cliente: Path = SAIDA_CLIENTE,
) -> tuple[Path, Path, pd.DataFrame]:
    entrada = Path(entrada or ARQUIVO_ENTRADA_PADRAO)
    PASTA_SAIDAS.mkdir(parents=True, exist_ok=True)

    df = ler_base_laboratorios(entrada)
    df_diretoria = df.copy()
    df_cliente = preparar_visao_danone(df)

    exportar_excel(df_diretoria, saida_diretoria, mascarado=False)
    exportar_excel(df_cliente, saida_cliente, mascarado=True)

    return saida_diretoria, saida_cliente, df_diretoria


if __name__ == "__main__":
    caminho_entrada = Path(sys.argv[1]) if len(sys.argv) > 1 else ARQUIVO_ENTRADA_PADRAO
    dir_path, cli_path, tabela = gerar_relatorios(caminho_entrada)

    print("Relatórios gerados com sucesso.\n")
    print(f"Entrada:  {caminho_entrada.resolve()}")
    print(f"Diretoria: {dir_path.resolve()}")
    print(f"Cliente Danone: {cli_path.resolve()}")
    print(f"\nLaboratórios incluídos ({len(tabela)}):")
    print(tabela[[COL_RANK, COL_LAB, COL_FAT26]].to_string(index=False))
