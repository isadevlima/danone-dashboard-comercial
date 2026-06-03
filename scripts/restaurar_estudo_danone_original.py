"""
Restaura o workbook no layout original (aba Ranking + blocos laterais)
a partir do CSV ESTUDO DANONE-01.csv — mesma estrutura da sua planilha manual.
"""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from danone.config import PASTA_DADOS

CSV_ORIGEM = PASTA_DADOS / "ESTUDO DANONE-01.csv"
XLSX_SAIDA = PASTA_DADOS / "ESTUDO DANONE-01.xlsx"

FMT_MOEDA = '"R$" #.##0,00'
FMT_PCT = "0,00%"

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_POS = PatternFill("solid", fgColor="E2EFDA")
FILL_NEG = PatternFill("solid", fgColor="FCE4D6")
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)

COLS_MOEDA = (2, 3, 7, 8)
COLS_PCT = (4, 9)


def _parse_moeda(valor) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto in ("-", "nan"):
        return None
    if "R$" not in texto and not re.search(r"\d,\d{2}$", texto):
        return None
    texto = texto.replace("R$", "").replace("\xa0", " ").strip()
    texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def _parse_percentual(valor) -> float | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    if isinstance(valor, (int, float)):
        v = float(valor)
        return v / 100 if abs(v) > 1.5 else v
    texto = str(valor).strip()
    if "%" not in texto and not re.match(r"^-?\d+,\d+$", texto):
        return None
    texto = texto.replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return float(texto) / 100.0
    except ValueError:
        return None


def _ler_csv() -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return pd.read_csv(CSV_ORIGEM, sep=";", header=None, encoding=enc)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Não foi possível ler {CSV_ORIGEM}")


def _formatar_celula_valor(cell, col: int) -> None:
    val = cell.value
    if val is None:
        return
    if col in COLS_MOEDA:
        n = _parse_moeda(val)
        if n is not None:
            cell.value = n
            cell.number_format = FMT_MOEDA
    elif col in COLS_PCT:
        p = _parse_percentual(val)
        if p is not None:
            cell.value = p
            cell.number_format = FMT_PCT


def _formatar_aba_ranking(ws) -> None:
    secao = None
    max_row = ws.max_row

    for r in range(1, max_row + 1):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                _formatar_celula_valor(cell, c)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(vertical="center")

        lab_a = ws.cell(row=r, column=1).value
        lab_f = ws.cell(row=r, column=6).value

        if str(lab_a or "").strip().lower() in ("laboratorio", "laboratório"):
            for c in range(1, 5):
                h = ws.cell(row=r, column=c)
                h.fill = FILL_HEADER
                h.font = FONT_HEADER
            for c in range(6, 10):
                h = ws.cell(row=r, column=c)
                h.fill = FILL_HEADER
                h.font = FONT_HEADER

        rotulo_f = str(lab_f or "").strip().lower()
        if rotulo_f == "impactos positivos":
            secao = "pos"
        elif rotulo_f == "impactos negativos":
            secao = "neg"
        elif rotulo_f.startswith("impactos"):
            secao = None

        if secao == "pos":
            for c in range(6, 11):
                ws.cell(row=r, column=c).fill = FILL_POS
        elif secao == "neg":
            for c in range(6, 11):
                ws.cell(row=r, column=c).fill = FILL_NEG

    larguras = {1: 38, 2: 20, 3: 20, 4: 14, 6: 38, 7: 20, 8: 20, 9: 14, 10: 22, 11: 10}
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 85


def _criar_aba_visao(wb, df: pd.DataFrame) -> None:
    nome = "VISAO_SOCIO_COLABORADORES_10351"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome)

    headers = ["Laboratorio", "Faturamento 2025", "Faturamento 2026", "Soma de Crescimento%"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    out_row = 2
    for _, row in df.iterrows():
        lab = row.iloc[0]
        if pd.isna(lab) or str(lab).strip().lower() in ("laboratorio", "laboratório", "nan", ""):
            continue
        f25, f26, pct = row.iloc[1], row.iloc[2], row.iloc[3]
        if pd.isna(f25) and pd.isna(f26):
            continue

        ws.cell(row=out_row, column=1, value=str(lab).strip())
        for c_idx, raw, parser, fmt in (
            (2, f25, _parse_moeda, FMT_MOEDA),
            (3, f26, _parse_moeda, FMT_MOEDA),
            (4, pct, _parse_percentual, FMT_PCT),
        ):
            parsed = parser(raw)
            cell = ws.cell(row=out_row, column=c_idx)
            if parsed is not None:
                cell.value = parsed
                cell.number_format = fmt
            elif not pd.isna(raw):
                cell.value = str(raw).strip()
        out_row += 1

    for col, w in zip(range(1, 5), (40, 22, 22, 16)):
        ws.column_dimensions[get_column_letter(col)].width = w
    if out_row > 2:
        ws.auto_filter.ref = f"A1:D{out_row - 1}"


def _criar_aba_analise(wb, df: pd.DataFrame) -> None:
    nome = "Analise Básica"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome)

    max_row = min(len(df), 30)
    for r in range(max_row):
        for c in range(5, min(11, df.shape[1])):
            val = df.iloc[r, c]
            if pd.isna(val):
                continue
            col_excel = c - 4
            cell = ws.cell(row=r + 1, column=col_excel, value=str(val).strip())
            _formatar_celula_valor(cell, col_excel)

    for col, w in zip(range(1, 8), (38, 22, 22, 16, 22, 12, 10)):
        ws.column_dimensions[get_column_letter(col)].width = w


def restaurar_workbook(
    csv_path: Path = CSV_ORIGEM,
    xlsx_path: Path = XLSX_SAIDA,
) -> Path:
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV não encontrado: {csv_path}")

    print("Lendo CSV...")
    df = _ler_csv()

    print("Gravando aba Ranking (pode levar 1–2 min)...")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Ranking", index=False, header=False)

    print("Aplicando formatação e demais abas...")
    wb = load_workbook(xlsx_path)
    _formatar_aba_ranking(wb["Ranking"])
    _criar_aba_visao(wb, df)
    _criar_aba_analise(wb, df)
    wb.save(xlsx_path)

    return xlsx_path.resolve()


if __name__ == "__main__":
    destino = restaurar_workbook()
    print(f"\nWorkbook restaurado: {destino}")
    print("Abas: Ranking | VISAO_SOCIO_COLABORADORES_10351 | Analise Básica")
    print("Abra ESTUDO DANONE-01.xlsx — layout como antes do script automático.")
